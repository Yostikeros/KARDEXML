from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from .forms import ClasificarBloqueForm, ClasificarDetalleForm, StockInicialForm
from .models import (
    Documento,
    DocumentoDetalle,
    EmpresaPrincipal,
    Entidad,
    MovimientoKardex,
    Producto,
    ProductoEquivalencia,
    TipoCambioSunat,
)
from .services.clasificacion import clasificar_detalle, excluir_detalle_kardex
from .services.dashboard import obtener_analisis_compras_ventas
from .services.kardex import (
    KardexError,
    confirmar_documento_kardex,
    editar_stock_inicial,
    registrar_stock_inicial,
    revertir_aprobacion_kardex,
    revertir_pre_kardex,
)
from .services.reportes import obtener_stock_actual
from .templatetags.kardex_format import fecha_corta, numero
from .services.xml_importer import XMLImportError, importar_xml


FACTURA_XML = b'''<?xml version="1.0" encoding="UTF-8"?>
<Invoice xmlns="urn:oasis:names:specification:ubl:schema:xsd:Invoice-2"
         xmlns:cac="urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2"
         xmlns:cbc="urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2">
  <cbc:UBLVersionID>2.1</cbc:UBLVersionID>
  <cbc:ID>F001-123</cbc:ID>
  <cbc:IssueDate>2026-06-24</cbc:IssueDate>
  <cbc:InvoiceTypeCode>01</cbc:InvoiceTypeCode>
  <cbc:DocumentCurrencyCode>PEN</cbc:DocumentCurrencyCode>
  <cac:AccountingSupplierParty>
    <cac:Party>
      <cac:PartyIdentification><cbc:ID schemeID="6">20111111111</cbc:ID></cac:PartyIdentification>
      <cac:PartyLegalEntity><cbc:RegistrationName>Proveedor SAC</cbc:RegistrationName></cac:PartyLegalEntity>
    </cac:Party>
  </cac:AccountingSupplierParty>
  <cac:AccountingCustomerParty>
    <cac:Party>
      <cac:PartyIdentification><cbc:ID schemeID="6">20999999999</cbc:ID></cac:PartyIdentification>
      <cac:PartyLegalEntity><cbc:RegistrationName>Empresa Principal SAC</cbc:RegistrationName></cac:PartyLegalEntity>
    </cac:Party>
  </cac:AccountingCustomerParty>
  <cac:LegalMonetaryTotal><cbc:PayableAmount currencyID="PEN">118.00</cbc:PayableAmount></cac:LegalMonetaryTotal>
  <cac:InvoiceLine>
    <cbc:ID>1</cbc:ID>
    <cbc:InvoicedQuantity unitCode="KG">10</cbc:InvoicedQuantity>
    <cbc:LineExtensionAmount currencyID="PEN">100.00</cbc:LineExtensionAmount>
    <cac:PricingReference><cac:AlternativeConditionPrice><cbc:PriceAmount currencyID="PEN">11.80</cbc:PriceAmount></cac:AlternativeConditionPrice></cac:PricingReference>
    <cac:TaxTotal><cbc:TaxAmount currencyID="PEN">18.00</cbc:TaxAmount></cac:TaxTotal>
    <cac:Item>
      <cbc:Description>Cafe pergamino</cbc:Description>
      <cac:SellersItemIdentification><cbc:ID>CAF-001</cbc:ID></cac:SellersItemIdentification>
    </cac:Item>
    <cac:Price><cbc:PriceAmount currencyID="PEN">10.00</cbc:PriceAmount></cac:Price>
  </cac:InvoiceLine>
</Invoice>'''


LIQUIDACION_XML = b'''<?xml version="1.0" encoding="UTF-8"?>
<Invoice xmlns="urn:oasis:names:specification:ubl:schema:xsd:Invoice-2"
         xmlns:cac="urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2"
         xmlns:cbc="urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2">
  <cbc:UBLVersionID>2.1</cbc:UBLVersionID>
  <cbc:ID>L001-9</cbc:ID>
  <cbc:IssueDate>2026-06-24</cbc:IssueDate>
  <cbc:InvoiceTypeCode>04</cbc:InvoiceTypeCode>
  <cbc:DocumentCurrencyCode>PEN</cbc:DocumentCurrencyCode>
  <cac:AccountingSupplierParty>
    <cac:Party>
      <cac:PartyIdentification><cbc:ID schemeID="6">20999999999</cbc:ID></cac:PartyIdentification>
      <cac:PartyLegalEntity><cbc:RegistrationName>Empresa Principal SAC</cbc:RegistrationName></cac:PartyLegalEntity>
    </cac:Party>
  </cac:AccountingSupplierParty>
  <cac:AccountingCustomerParty>
    <cac:Party>
      <cac:PartyIdentification><cbc:ID schemeID="6">20888888888</cbc:ID></cac:PartyIdentification>
      <cac:PartyLegalEntity><cbc:RegistrationName>Productor Cafe</cbc:RegistrationName></cac:PartyLegalEntity>
    </cac:Party>
  </cac:AccountingCustomerParty>
  <cac:LegalMonetaryTotal><cbc:PayableAmount currencyID="PEN">118.00</cbc:PayableAmount></cac:LegalMonetaryTotal>
  <cac:InvoiceLine>
    <cbc:ID>1</cbc:ID>
    <cbc:InvoicedQuantity unitCode="KG">10</cbc:InvoicedQuantity>
    <cbc:LineExtensionAmount currencyID="PEN">100.00</cbc:LineExtensionAmount>
    <cac:TaxTotal><cbc:TaxAmount currencyID="PEN">18.00</cbc:TaxAmount></cac:TaxTotal>
    <cac:Item>
      <cbc:Description>Cafe pergamino</cbc:Description>
      <cac:SellersItemIdentification><cbc:ID>CAF-001</cbc:ID></cac:SellersItemIdentification>
    </cac:Item>
    <cac:Price><cbc:PriceAmount currencyID="PEN">10.00</cbc:PriceAmount></cac:Price>
  </cac:InvoiceLine>
</Invoice>'''
VENTA_USD_LBR_XML = b'''<?xml version="1.0" encoding="UTF-8"?>
<Invoice xmlns="urn:oasis:names:specification:ubl:schema:xsd:Invoice-2"
         xmlns:cac="urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2"
         xmlns:cbc="urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2">
  <cbc:UBLVersionID>2.1</cbc:UBLVersionID>
  <cbc:ID>F002-456</cbc:ID>
  <cbc:IssueDate>2026-06-24</cbc:IssueDate>
  <cbc:InvoiceTypeCode>01</cbc:InvoiceTypeCode>
  <cbc:DocumentCurrencyCode listAgencyName="United Nations Economic Commission for Europe" listID="ISO 4217 Alpha" listName="Currency">USD</cbc:DocumentCurrencyCode>
  <cac:AccountingSupplierParty>
    <cac:Party>
      <cac:PartyIdentification><cbc:ID schemeID="6">20999999999</cbc:ID></cac:PartyIdentification>
      <cac:PartyLegalEntity><cbc:RegistrationName>Empresa Principal SAC</cbc:RegistrationName></cac:PartyLegalEntity>
    </cac:Party>
  </cac:AccountingSupplierParty>
  <cac:AccountingCustomerParty>
    <cac:Party>
      <cac:PartyIdentification><cbc:ID schemeID="6">20111111111</cbc:ID></cac:PartyIdentification>
      <cac:PartyLegalEntity><cbc:RegistrationName>Cliente Exterior SAC</cbc:RegistrationName></cac:PartyLegalEntity>
    </cac:Party>
  </cac:AccountingCustomerParty>
  <cac:LegalMonetaryTotal><cbc:PayableAmount currencyID="USD">118.00</cbc:PayableAmount></cac:LegalMonetaryTotal>
  <cac:InvoiceLine>
    <cbc:ID>1</cbc:ID>
    <cbc:InvoicedQuantity unitCode="LBR" unitCodeListAgencyName="United Nations Economic Commission for Europe" unitCodeListID="UN/ECE rec 20">43353.459</cbc:InvoicedQuantity>
    <cbc:LineExtensionAmount currencyID="USD">100.00</cbc:LineExtensionAmount>
    <cac:PricingReference><cac:AlternativeConditionPrice><cbc:PriceAmount currencyID="USD">11.80</cbc:PriceAmount></cac:AlternativeConditionPrice></cac:PricingReference>
    <cac:TaxTotal><cbc:TaxAmount currencyID="USD">18.00</cbc:TaxAmount></cac:TaxTotal>
    <cac:Item>
      <cbc:Description>Cafe pergamino</cbc:Description>
      <cac:SellersItemIdentification><cbc:ID>CAF-001</cbc:ID></cac:SellersItemIdentification>
    </cac:Item>
    <cac:Price><cbc:PriceAmount currencyID="USD">10.00</cbc:PriceAmount></cac:Price>
  </cac:InvoiceLine>
</Invoice>'''

LIQUIDACION_DNI_XML = LIQUIDACION_XML.replace(
    b'<cac:PartyIdentification><cbc:ID schemeID="6">20888888888</cbc:ID></cac:PartyIdentification>',
    b'<cac:PartyIdentification><cbc:ID>27833207</cbc:ID></cac:PartyIdentification>',
).replace(
    b'<cac:PartyLegalEntity><cbc:RegistrationName>Productor Cafe</cbc:RegistrationName></cac:PartyLegalEntity>',
    b'<cac:PartyLegalEntity><cbc:RegistrationName>Productor DNI</cbc:RegistrationName></cac:PartyLegalEntity>',
)

LIQUIDACION_NIU_KG_XML = LIQUIDACION_XML.replace(
    b'<cbc:InvoicedQuantity unitCode="KG">10</cbc:InvoicedQuantity>',
    b'<cbc:InvoicedQuantity unitCode="NIU">1</cbc:InvoicedQuantity>',
).replace(
    b'<cbc:Description>Cafe pergamino</cbc:Description>',
    b'<cbc:Description>Cafe pergamino productor 250 KG</cbc:Description>',
).replace(
    b'<cac:Price><cbc:PriceAmount currencyID="PEN">10.00</cbc:PriceAmount></cac:Price>',
    b'<cac:Price><cbc:PriceAmount currencyID="PEN">0.00</cbc:PriceAmount></cac:Price>',
)

class XMLImporterTests(TestCase):
    def setUp(self):
        EmpresaPrincipal.objects.create(
            ruc='20999999999',
            razon_social='Empresa Principal SAC',
        )

    def test_importa_factura_de_compra_y_crea_proveedor(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')

        self.assertTrue(result.ok)
        self.assertEqual(result.documento.tipo_operacion, Documento.COMPRA)
        self.assertEqual(result.documento.proveedor.numero_documento, '20111111111')
        self.assertEqual(result.pendientes_clasificacion, 1)
        self.assertTrue(Entidad.objects.get(numero_documento='20111111111').es_proveedor)

    def test_pendientes_muestra_detalles_de_producto_xml(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        user = User.objects.create_user(username='admin', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:documentos_pendientes'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Cafe pergamino')
        self.assertContains(response, 'CAF-001')
        self.assertContains(response, reverse('kardex:comprobante_documento', args=[result.documento.id]))

    def test_comprobante_documento_muestra_factura_representada(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        user = User.objects.create_user(username='admin', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:comprobante_documento', args=[result.documento.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Proveedor SAC')
        self.assertContains(response, 'Empresa Principal SAC')
        self.assertContains(response, 'F001-123')
        self.assertContains(response, 'Cafe pergamino')

    def test_importa_factura_venta_usd_lbr_convierte_moneda_y_unidad(self):
        TipoCambioSunat.objects.create(
            mes='junio',
            anio=2026,
            dia=24,
            compra='3.600000',
            venta='3.700000',
        )

        result = importar_xml(VENTA_USD_LBR_XML, filename='venta-usd-lbr.xml')

        detalle = result.documento.detalles.first()
        self.assertTrue(result.ok)
        self.assertEqual(result.documento.tipo_operacion, Documento.VENTA)
        self.assertEqual(result.documento.moneda, 'USD')
        self.assertEqual(result.documento.tipo_cambio, Decimal('3.700000'))
        self.assertEqual(result.documento.total, Decimal('436.600'))
        self.assertEqual(detalle.unidad_medida_xml, 'LBR')
        self.assertEqual(detalle.factor_conversion, Decimal('0.453592'))
        self.assertEqual(detalle.cantidad_base.quantize(Decimal('0.01')), Decimal('19664.78'))
        self.assertEqual(detalle.subtotal, Decimal('370.000'))
        self.assertEqual(detalle.igv, Decimal('66.600'))

    def test_comprobante_usd_lbr_muestra_moneda_y_unidad_operativa(self):
        TipoCambioSunat.objects.create(
            mes='junio',
            anio=2026,
            dia=24,
            compra='3.600000',
            venta='3.700000',
        )
        result = importar_xml(VENTA_USD_LBR_XML, filename='venta-usd-lbr.xml')
        user = User.objects.create_user(username='visor-usd', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:comprobante_documento', args=[result.documento.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'USD')
        self.assertNotContains(response, 'ISO')
        self.assertContains(response, 'KG')
        self.assertNotContains(response, 'NIU')
        self.assertContains(response, '19,664.78')

    def test_importa_kgm_como_kg(self):
        xml = VENTA_USD_LBR_XML.replace(
            b'<cbc:InvoicedQuantity unitCode="LBR" unitCodeListAgencyName="United Nations Economic Commission for Europe" unitCodeListID="UN/ECE rec 20">43353.459</cbc:InvoicedQuantity>',
            b'<cbc:InvoicedQuantity unitCode="KGM" unitCodeListAgencyName="United Nations Economic Commission for Europe" unitCodeListID="UN/ECE rec 20">17733.00</cbc:InvoicedQuantity>',
        )
        TipoCambioSunat.objects.create(
            mes='junio',
            anio=2026,
            dia=24,
            compra='3.600000',
            venta='3.700000',
        )

        result = importar_xml(xml, filename='venta-kgm.xml')

        detalle = result.documento.detalles.first()
        self.assertEqual(detalle.unidad_medida_xml, 'KG')
        self.assertEqual(detalle.cantidad, Decimal('17733.000000'))
        self.assertEqual(detalle.cantidad_base, Decimal('17733.000000'))

    def test_importa_niu_con_kg_en_descripcion_como_kg(self):
        xml = VENTA_USD_LBR_XML.replace(
            b'<cbc:InvoicedQuantity unitCode="LBR" unitCodeListAgencyName="United Nations Economic Commission for Europe" unitCodeListID="UN/ECE rec 20">43353.459</cbc:InvoicedQuantity>',
            b'<cbc:InvoicedQuantity unitCode="NIU">0</cbc:InvoicedQuantity>',
        ).replace(
            b'<cbc:Description>Cafe pergamino</cbc:Description>',
            b'<cbc:Description>EQUIVALENT TO 285 BAGS OF 69 KG NET OR 19665 KG NET OF WASHED GREEN ARABICA COFFEE</cbc:Description>',
        )
        TipoCambioSunat.objects.create(
            mes='junio',
            anio=2026,
            dia=24,
            compra='3.600000',
            venta='3.700000',
        )

        result = importar_xml(xml, filename='venta-niu-kg.xml')

        detalle = result.documento.detalles.first()
        self.assertEqual(detalle.unidad_medida_xml, 'KG')
        self.assertEqual(detalle.cantidad, Decimal('19665'))
        self.assertEqual(detalle.cantidad_base, Decimal('19665.000000'))

    def test_importa_usd_requiere_tipo_cambio_sunat(self):
        with self.assertRaises(XMLImportError):
            importar_xml(VENTA_USD_LBR_XML, filename='venta-usd-lbr.xml')

    def test_importa_liquidacion_de_compra_con_receptor_como_proveedor(self):
        result = importar_xml(LIQUIDACION_XML, filename='liquidacion.xml')

        self.assertTrue(result.ok)
        self.assertEqual(result.documento.tipo_operacion, Documento.COMPRA)
        self.assertEqual(result.documento.proveedor.numero_documento, '20888888888')
        self.assertTrue(Entidad.objects.get(numero_documento='20888888888').es_proveedor)

    def test_importa_liquidacion_de_compra_con_receptor_dni_sin_scheme(self):
        result = importar_xml(LIQUIDACION_DNI_XML, filename='liquidacion-dni.xml')

        entidad = Entidad.objects.get(numero_documento='27833207')
        self.assertTrue(result.ok)
        self.assertEqual(result.documento.tipo_documento, Documento.LIQUIDACION_COMPRA)
        self.assertEqual(result.documento.tipo_operacion, Documento.COMPRA)
        self.assertEqual(result.documento.proveedor, entidad)
        self.assertEqual(entidad.tipo_documento_identidad, Entidad.DNI)
        self.assertTrue(entidad.es_proveedor)

    def test_importa_liquidacion_niu_con_kilos_y_precio_derivado(self):
        result = importar_xml(LIQUIDACION_NIU_KG_XML, filename='liquidacion-niu-kg.xml')

        detalle = result.documento.detalles.first()
        self.assertTrue(result.ok)
        self.assertEqual(detalle.unidad_medida_xml, 'KG')
        self.assertEqual(detalle.cantidad, Decimal('250.000000'))
        self.assertEqual(detalle.cantidad_base, Decimal('250.000000'))
        self.assertEqual(detalle.valor_unitario.quantize(Decimal('0.000001')), Decimal('0.400000'))
        self.assertEqual(detalle.precio_unitario.quantize(Decimal('0.000001')), Decimal('0.472000'))

    def test_importacion_corrige_entidad_existente_con_dni_como_ruc(self):
        Entidad.objects.create(
            tipo_documento_identidad=Entidad.RUC,
            numero_documento='27833207',
            razon_social='Productor previo',
        )

        importar_xml(LIQUIDACION_DNI_XML, filename='liquidacion-dni.xml')

        entidad = Entidad.objects.get(numero_documento='27833207')
        self.assertEqual(entidad.tipo_documento_identidad, Entidad.DNI)
        self.assertTrue(entidad.es_proveedor)

    def test_comprobante_muestra_tipo_documento_sunat(self):
        result = importar_xml(LIQUIDACION_XML, filename='liquidacion.xml')
        user = User.objects.create_user(username='admin', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:comprobante_documento', args=[result.documento.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Liquidacion de compra')
        self.assertNotContains(response, '<div class="fs-4 fw-bold text-uppercase">Otro</div>', html=True)

    def test_clasificar_documento_muestra_tipo_documento_sunat(self):
        result = importar_xml(LIQUIDACION_XML, filename='liquidacion.xml')
        user = User.objects.create_user(username='admin2', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:clasificar_documento', args=[result.documento.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Liquidacion de compra')

    def test_rechaza_xml_duplicado_por_hash(self):
        importar_xml(FACTURA_XML, filename='factura.xml')

        with self.assertRaises(XMLImportError):
            importar_xml(FACTURA_XML, filename='factura.xml')

    def test_documentos_lista_muestra_compra_para_liquidacion_antigua(self):
        result = importar_xml(LIQUIDACION_XML, filename='liquidacion.xml')
        result.documento.tipo_operacion = Documento.OTRO
        result.documento.save(update_fields=['tipo_operacion'])
        user = User.objects.create_user(username='docs', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:documentos_lista'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<td>Compra</td>', html=True)

    def test_documentos_lista_permite_ordenar_por_tipo_documento(self):
        importar_xml(LIQUIDACION_XML, filename='liquidacion.xml')
        importar_xml(FACTURA_XML, filename='factura.xml')
        user = User.objects.create_user(username='docs-orden', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:documentos_lista'), {'orden': 'tipo_documento'})
        content = response.content.decode()

        self.assertEqual(response.status_code, 200)
        self.assertLess(content.index('01-F001-123'), content.index('04-L001-9'))
        self.assertContains(response, '<option value="tipo_documento" selected>Tipo de documento</option>', html=True)

    def test_documentos_lista_filtra_por_tipo_documento(self):
        importar_xml(LIQUIDACION_XML, filename='liquidacion.xml')
        importar_xml(FACTURA_XML, filename='factura.xml')
        user = User.objects.create_user(username='docs-filtro-tipo', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(
            reverse('kardex:documentos_lista'),
            {'tipo_documento': Documento.LIQUIDACION_COMPRA, 'orden': 'tipo_documento'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '04-L001-9')
        self.assertNotContains(response, '01-F001-123')
        self.assertContains(response, '<option value="04" selected>Liquidacion de compra</option>', html=True)

    def test_documentos_lista_filtra_por_serie_o_numero(self):
        importar_xml(LIQUIDACION_XML, filename='liquidacion.xml')
        importar_xml(FACTURA_XML, filename='factura.xml')
        user = User.objects.create_user(username='docs-filtro-doc', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:documentos_lista'), {'documento': 'F001'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '01-F001-123')
        self.assertNotContains(response, '04-L001-9')

    def test_documentos_lista_muestra_controles_de_clasificacion_masiva(self):
        importar_xml(FACTURA_XML, filename='factura.xml')
        user = User.objects.create_user(username='docs-masivo', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:documentos_lista'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Clasificar lote')
        self.assertContains(response, 'data-select-documentos="all"')
        self.assertContains(response, 'name="documento_ids"')

    def test_documentos_lista_clasifica_documentos_por_bloque(self):
        factura = importar_xml(FACTURA_XML, filename='factura.xml').documento
        liquidacion = importar_xml(LIQUIDACION_XML, filename='liquidacion.xml').documento
        producto = Producto.objects.create(
            codigo_interno='MP-DOCS',
            nombre='Cafe desde documentos',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        user = User.objects.create_user(username='docs-bloque', password='admin12345')
        self.client.force_login(user)

        response = self.client.post(
            reverse('kardex:clasificar_documentos_bloque'),
            {
                'documentos': [str(factura.id), str(liquidacion.id)],
                'producto': str(producto.id),
                'factor_conversion': '1',
                'guardar_equivalencia': 'on',
            },
        )

        self.assertRedirects(response, reverse('kardex:documentos_lista'))
        factura.refresh_from_db()
        liquidacion.refresh_from_db()
        self.assertEqual(factura.estado, Documento.PRE_KARDEX)
        self.assertEqual(liquidacion.estado, Documento.PRE_KARDEX)
        self.assertFalse(
            DocumentoDetalle.objects.filter(
                documento__in=[factura, liquidacion],
                estado_clasificacion=DocumentoDetalle.PENDIENTE,
            ).exists()
        )
        self.assertEqual(
            DocumentoDetalle.objects.filter(
                documento__in=[factura, liquidacion],
                producto=producto,
                estado_clasificacion=DocumentoDetalle.CLASIFICADO,
            ).count(),
            2,
        )

    def test_clasifica_detalles_por_bloque(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        documento = result.documento
        detalle_original = documento.detalles.first()
        detalle_extra = DocumentoDetalle.objects.create(
            documento=documento,
            codigo_producto_xml='CAF-002',
            descripcion_xml='Cafe bola',
            unidad_medida_xml='KG',
            cantidad=5,
            valor_unitario=10,
            precio_unitario=10,
            subtotal=50,
            igv=9,
            total=59,
        )
        producto = Producto.objects.create(
            codigo_interno='MP-BLOQUE',
            nombre='Cafe clasificado por bloque',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        user = User.objects.create_user(username='bloque', password='admin12345')
        self.client.force_login(user)

        response = self.client.post(
            reverse('kardex:clasificar_bloque', args=[documento.id]),
            {
                'detalles': [str(detalle_original.id), str(detalle_extra.id)],
                'producto': str(producto.id),
                'factor_conversion': '1',
                'guardar_equivalencia': 'on',
            },
        )

        self.assertRedirects(response, reverse('kardex:clasificar_documento', args=[documento.id]))
        documento.refresh_from_db()
        detalle_original.refresh_from_db()
        detalle_extra.refresh_from_db()
        self.assertEqual(documento.estado, Documento.PRE_KARDEX)
        self.assertEqual(detalle_original.producto, producto)
        self.assertEqual(detalle_extra.producto, producto)
        self.assertEqual(detalle_original.estado_clasificacion, DocumentoDetalle.CLASIFICADO)
        self.assertEqual(detalle_extra.estado_clasificacion, DocumentoDetalle.CLASIFICADO)

    def test_clasificar_documentos_por_bloque_sin_seleccion_no_genera_500(self):
        producto = Producto.objects.create(
            codigo_interno='MP-SIN-DOCS',
            nombre='Cafe sin documentos seleccionados',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        user = User.objects.create_user(username='sin-docs', password='admin12345')
        self.client.force_login(user)

        response = self.client.post(
            reverse('kardex:clasificar_documentos_bloque'),
            {
                'producto': str(producto.id),
                'factor_conversion': '1',
                'guardar_equivalencia': 'on',
            },
        )

        self.assertRedirects(response, reverse('kardex:documentos_lista'))

    def test_clasificar_detalles_por_bloque_sin_seleccion_no_genera_500(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        producto = Producto.objects.create(
            codigo_interno='MP-SIN-DET',
            nombre='Cafe sin detalles seleccionados',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        user = User.objects.create_user(username='sin-detalles', password='admin12345')
        self.client.force_login(user)

        response = self.client.post(
            reverse('kardex:clasificar_bloque', args=[result.documento.id]),
            {
                'producto': str(producto.id),
                'factor_conversion': '1',
                'guardar_equivalencia': 'on',
            },
        )

        self.assertRedirects(response, reverse('kardex:clasificar_documento', args=[result.documento.id]))

    def test_clasifica_detalle_crea_equivalencia_y_pasa_a_pre_kardex(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        producto = Producto.objects.create(
            codigo_interno='MP-CAFE',
            nombre='Cafe pergamino',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        detalle = result.documento.detalles.first()

        clasificar_detalle(
            detalle=detalle,
            producto=producto,
            factor_conversion='1',
            guardar_equivalencia=True,
        )

        result.documento.refresh_from_db()
        detalle.refresh_from_db()
        self.assertEqual(result.documento.estado, Documento.PRE_KARDEX)
        self.assertEqual(detalle.estado_clasificacion, DocumentoDetalle.CLASIFICADO)
        self.assertTrue(
            ProductoEquivalencia.objects.filter(
                codigo_producto_xml='CAF-001',
                descripcion_xml='Cafe pergamino',
                producto=producto,
            ).exists()
        )

    def test_excluye_detalle_y_permite_pasar_a_pre_kardex(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        detalle = result.documento.detalles.first()

        excluir_detalle_kardex(detalle)

        result.documento.refresh_from_db()
        detalle.refresh_from_db()
        self.assertEqual(result.documento.estado, Documento.PRE_KARDEX)
        self.assertEqual(detalle.estado_clasificacion, DocumentoDetalle.NO_APLICA)
        self.assertFalse(detalle.afecta_kardex)
        self.assertEqual(detalle.cantidad_base, 0)

    def test_aprobar_documento_con_detalle_excluido_no_genera_movimiento(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        excluir_detalle_kardex(result.documento.detalles.first())

        movimientos = confirmar_documento_kardex(result.documento)

        result.documento.refresh_from_db()
        self.assertEqual(result.documento.estado, Documento.CONFIRMADO)
        self.assertEqual(movimientos, [])
        self.assertFalse(MovimientoKardex.objects.filter(documento_origen=result.documento).exists())

    def test_pre_kardex_aprueba_documentos_por_bloque(self):
        factura = importar_xml(FACTURA_XML, filename='factura.xml').documento
        liquidacion = importar_xml(LIQUIDACION_XML, filename='liquidacion.xml').documento
        producto = Producto.objects.create(
            codigo_interno='MP-PRE-BLOQUE',
            nombre='Cafe Pre-Kardex bloque',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        for documento in [factura, liquidacion]:
            for detalle in documento.detalles.all():
                clasificar_detalle(detalle, producto, '1', guardar_equivalencia=False)
        user = User.objects.create_user(username='pre-bloque', password='admin12345')
        self.client.force_login(user)

        response = self.client.post(
            reverse('kardex:confirmar_kardex_bloque'),
            {'documentos': [str(factura.id), str(liquidacion.id)]},
        )

        self.assertRedirects(response, reverse('kardex:pre_kardex_lista'))
        factura.refresh_from_db()
        liquidacion.refresh_from_db()
        self.assertEqual(factura.estado, Documento.CONFIRMADO)
        self.assertEqual(liquidacion.estado, Documento.CONFIRMADO)
        self.assertEqual(
            MovimientoKardex.objects.filter(documento_origen__in=[factura, liquidacion]).count(),
            2,
        )

    def test_bandeja_pre_kardex_muestra_documento_para_aprobar(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        producto = Producto.objects.create(
            codigo_interno='MP-CAFE',
            nombre='Cafe pergamino',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        clasificar_detalle(
            detalle=result.documento.detalles.first(),
            producto=producto,
            factor_conversion='1',
            guardar_equivalencia=False,
        )
        user = User.objects.create_user(username='admin', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:pre_kardex_lista'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Aprobar Pre-Kardex')
        self.assertContains(response, 'Cafe pergamino')

    def test_confirma_kardex_de_compra_con_promedio_movil(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        producto = Producto.objects.create(
            codigo_interno='MP-CAFE',
            nombre='Cafe pergamino',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        clasificar_detalle(
            detalle=result.documento.detalles.first(),
            producto=producto,
            factor_conversion='1',
            guardar_equivalencia=False,
        )

        movimientos = confirmar_documento_kardex(result.documento)

        result.documento.refresh_from_db()
        self.assertEqual(result.documento.estado, Documento.CONFIRMADO)
        self.assertEqual(len(movimientos), 1)
        movimiento = MovimientoKardex.objects.get(documento_origen=result.documento)
        self.assertEqual(movimiento.tipo_movimiento, MovimientoKardex.ENTRADA)
        self.assertEqual(movimiento.cantidad_entrada, 10)
        self.assertEqual(movimiento.costo_total_entrada, 100)
        self.assertEqual(movimiento.stock_cantidad, 10)
        self.assertEqual(movimiento.stock_costo_unitario_promedio, 10)

    def test_no_permite_confirmar_dos_veces(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        producto = Producto.objects.create(
            codigo_interno='MP-CAFE',
            nombre='Cafe pergamino',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        clasificar_detalle(
            detalle=result.documento.detalles.first(),
            producto=producto,
            factor_conversion='1',
            guardar_equivalencia=False,
        )
        confirmar_documento_kardex(result.documento)

        with self.assertRaises(KardexError):
            confirmar_documento_kardex(result.documento)

    def test_revierte_pre_kardex_a_pendiente_y_limpia_detalles(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        producto = Producto.objects.create(
            codigo_interno='MP-REV-PRE',
            nombre='Cafe revierte Pre-Kardex',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        detalle = result.documento.detalles.first()
        clasificar_detalle(detalle, producto, '1', guardar_equivalencia=False)

        revertir_pre_kardex(result.documento)

        result.documento.refresh_from_db()
        detalle.refresh_from_db()
        self.assertEqual(result.documento.estado, Documento.PENDIENTE_CLASIFICACION)
        self.assertEqual(detalle.estado_clasificacion, DocumentoDetalle.PENDIENTE)
        self.assertIsNone(detalle.producto)
        self.assertFalse(detalle.afecta_kardex)
        self.assertEqual(detalle.cantidad_base, 0)

    def test_quita_documento_de_pre_kardex_desde_bandeja_sin_eliminar_xml(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        producto = Producto.objects.create(
            codigo_interno='MP-QUITAR-PRE',
            nombre='Cafe quitar Pre-Kardex',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        clasificar_detalle(
            detalle=result.documento.detalles.first(),
            producto=producto,
            factor_conversion='1',
            guardar_equivalencia=False,
        )
        user = User.objects.create_user(username='quitar-pre', password='admin12345')
        self.client.force_login(user)

        response = self.client.post(
            reverse('kardex:revertir_pre_kardex', args=[result.documento.id]),
            {'next': 'pre_kardex_lista'},
        )

        self.assertRedirects(response, reverse('kardex:pre_kardex_lista'))
        result.documento.refresh_from_db()
        self.assertEqual(result.documento.estado, Documento.PENDIENTE_CLASIFICACION)
        self.assertTrue(Documento.objects.filter(pk=result.documento.pk).exists())
        self.assertTrue(DocumentoDetalle.objects.filter(documento=result.documento).exists())

    def test_revierte_aprobacion_y_vuelve_a_pre_kardex(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        producto = Producto.objects.create(
            codigo_interno='MP-REV-APR',
            nombre='Cafe revierte aprobacion',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        clasificar_detalle(
            detalle=result.documento.detalles.first(),
            producto=producto,
            factor_conversion='1',
            guardar_equivalencia=False,
        )
        confirmar_documento_kardex(result.documento)

        revertir_aprobacion_kardex(result.documento)

        result.documento.refresh_from_db()
        self.assertEqual(result.documento.estado, Documento.PRE_KARDEX)
        self.assertFalse(MovimientoKardex.objects.filter(documento_origen=result.documento).exists())
        self.assertFalse(MovimientoKardex.objects.filter(producto=producto).exists())

    def test_no_revierte_aprobacion_si_producto_tiene_movimientos_posteriores(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        producto = Producto.objects.create(
            codigo_interno='MP-REV-BLOCK',
            nombre='Cafe revierte bloqueado',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        clasificar_detalle(
            detalle=result.documento.detalles.first(),
            producto=producto,
            factor_conversion='1',
            guardar_equivalencia=False,
        )
        confirmar_documento_kardex(result.documento)
        MovimientoKardex.objects.create(
            fecha='2026-06-25',
            producto=producto,
            tipo_movimiento=MovimientoKardex.AJUSTE_ENTRADA,
            cantidad_entrada=1,
            costo_unitario_entrada=10,
            costo_total_entrada=10,
            stock_cantidad=11,
            stock_costo_unitario_promedio=10,
            stock_costo_total=110,
        )

        with self.assertRaises(KardexError):
            revertir_aprobacion_kardex(result.documento)

        result.documento.refresh_from_db()
        self.assertEqual(result.documento.estado, Documento.CONFIRMADO)
        self.assertTrue(MovimientoKardex.objects.filter(documento_origen=result.documento).exists())

    def test_reporte_stock_actual_usa_ultimo_saldo(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        producto = Producto.objects.create(
            codigo_interno='MP-CAFE',
            nombre='Cafe pergamino',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        clasificar_detalle(
            detalle=result.documento.detalles.first(),
            producto=producto,
            factor_conversion='1',
            guardar_equivalencia=False,
        )
        confirmar_documento_kardex(result.documento)

        item = next(item for item in obtener_stock_actual() if item.producto == producto)
        self.assertEqual(item.cantidad, 10)
        self.assertEqual(item.costo_promedio, 10)
        self.assertEqual(item.costo_total, 100)

    def test_paginas_de_reportes_cargan(self):
        user = User.objects.create_user(username='admin', password='admin12345')
        self.client.force_login(user)

        urls = [
            reverse('kardex:reportes_index'),
            reverse('kardex:reporte_stock_actual'),
            reverse('kardex:reporte_kardex_producto'),
            reverse('kardex:reporte_kardex_sunat_producto'),
            reverse('kardex:reporte_documentos_importados'),
            reverse('kardex:reporte_movimientos_documento'),
        ]
        for url in urls:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200, url)

    def test_reportes_exportan_excel(self):
        user = User.objects.create_user(username='excel-reportes', password='admin12345')
        self.client.force_login(user)

        urls = [
            reverse('kardex:reporte_stock_actual'),
            reverse('kardex:reporte_kardex_producto'),
            reverse('kardex:reporte_documentos_importados'),
            reverse('kardex:reporte_movimientos_documento'),
        ]
        for url in urls:
            response = self.client.get(url, {'export': 'excel'})
            self.assertEqual(response.status_code, 200, url)
            self.assertEqual(
                response['Content-Type'],
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            self.assertTrue(response.content.startswith(b'PK'))

    def test_reporte_kardex_sunat_producto_muestra_formato_13_1(self):
        result = importar_xml(FACTURA_XML, filename='factura.xml')
        producto = Producto.objects.create(
            codigo_interno='MP-SUNAT',
            nombre='Cafe SUNAT',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        clasificar_detalle(
            detalle=result.documento.detalles.first(),
            producto=producto,
            factor_conversion='1',
            guardar_equivalencia=False,
        )
        confirmar_documento_kardex(result.documento)
        user = User.objects.create_user(username='sunat-report', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(
            reverse('kardex:reporte_kardex_sunat_producto'),
            {
                'producto': producto.id,
                'fecha_inicio': '2026-06-01',
                'fecha_fin': '2026-06-30',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'FORMATO 13.1')
        self.assertContains(response, 'Empresa Principal SAC')
        self.assertContains(response, 'MP-SUNAT')
        self.assertContains(response, 'Cafe SUNAT')
        self.assertContains(response, '02 - Compra')
        self.assertContains(response, 'F001')
        self.assertContains(response, '123')
        self.assertContains(response, '10.00')
        self.assertContains(response, '100.00')

        export_response = self.client.get(
            reverse('kardex:reporte_kardex_sunat_producto'),
            {
                'producto': producto.id,
                'fecha_inicio': '2026-06-01',
                'fecha_fin': '2026-06-30',
                'export': 'excel',
            },
        )
        workbook = load_workbook(BytesIO(export_response.content))
        worksheet = workbook.active
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(worksheet.title, 'Formato 13.1')
        values = [cell.value for row in worksheet.iter_rows() for cell in row]
        self.assertIn('13.1 Registro de Inventario Permanente Valorizado', values)
        self.assertIn('02 - Compra', values)
        self.assertIn('F001', values)

    def test_registra_stock_inicial_como_ajuste_entrada(self):
        producto = Producto.objects.create(
            codigo_interno='MP-CAFE',
            nombre='Cafe pergamino',
            tipo_producto=Producto.MATERIA_PRIMA,
        )

        movimiento = registrar_stock_inicial(
            producto=producto,
            fecha='2026-06-01',
            cantidad=100,
            costo_unitario=12,
        )

        self.assertEqual(movimiento.tipo_movimiento, MovimientoKardex.AJUSTE_ENTRADA)
        self.assertEqual(movimiento.cantidad_entrada, 100)
        self.assertEqual(movimiento.costo_unitario_entrada, 12)
        self.assertEqual(movimiento.costo_total_entrada, 1200)
        self.assertEqual(movimiento.stock_cantidad, 100)
        self.assertEqual(movimiento.stock_costo_unitario_promedio, 12)
        self.assertEqual(movimiento.stock_costo_total, 1200)

    def test_no_permite_stock_inicial_si_producto_ya_tiene_movimientos(self):
        producto = Producto.objects.create(
            codigo_interno='MP-CAFE',
            nombre='Cafe pergamino',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        registrar_stock_inicial(
            producto=producto,
            fecha='2026-06-01',
            cantidad=100,
            costo_unitario=12,
        )

        with self.assertRaises(KardexError):
            registrar_stock_inicial(
                producto=producto,
                fecha='2026-06-02',
                cantidad=50,
                costo_unitario=13,
            )

    def test_pagina_stock_inicial_carga(self):
        user = User.objects.create_user(username='admin', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:stock_inicial'))

        self.assertEqual(response.status_code, 200)

    def test_edita_stock_inicial_si_no_hay_movimientos_posteriores(self):
        producto = Producto.objects.create(
            codigo_interno='MP-CAFE',
            nombre='Cafe pergamino',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        movimiento = registrar_stock_inicial(
            producto=producto,
            fecha='2026-06-01',
            cantidad=100,
            costo_unitario=12,
        )

        editar_stock_inicial(
            movimiento=movimiento,
            fecha='2026-06-02',
            cantidad=80,
            costo_unitario=11,
            observacion='Correccion de inventario inicial.',
        )

        movimiento.refresh_from_db()
        self.assertEqual(movimiento.cantidad_entrada, 80)
        self.assertEqual(movimiento.costo_unitario_entrada, 11)
        self.assertEqual(movimiento.costo_total_entrada, 880)
        self.assertEqual(movimiento.stock_cantidad, 80)
        self.assertEqual(movimiento.stock_costo_total, 880)

    def test_no_edita_stock_inicial_si_ya_hay_otro_movimiento(self):
        producto = Producto.objects.create(
            codigo_interno='MP-CAFE',
            nombre='Cafe pergamino',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        movimiento = registrar_stock_inicial(
            producto=producto,
            fecha='2026-06-01',
            cantidad=100,
            costo_unitario=12,
        )
        MovimientoKardex.objects.create(
            fecha='2026-06-03',
            producto=producto,
            tipo_movimiento=MovimientoKardex.AJUSTE_ENTRADA,
            cantidad_entrada=1,
            costo_unitario_entrada=12,
            costo_total_entrada=12,
            stock_cantidad=101,
            stock_costo_unitario_promedio=12,
            stock_costo_total=1212,
        )

        with self.assertRaises(KardexError):
            editar_stock_inicial(
                movimiento=movimiento,
                fecha='2026-06-02',
                cantidad=80,
                costo_unitario=11,
            )

    def test_pagina_editar_stock_inicial_carga(self):
        user = User.objects.create_user(username='admin', password='admin12345')
        self.client.force_login(user)
        producto = Producto.objects.create(
            codigo_interno='MP-CAFE',
            nombre='Cafe pergamino',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        movimiento = registrar_stock_inicial(
            producto=producto,
            fecha='2026-06-01',
            cantidad=100,
            costo_unitario=12,
        )

        response = self.client.get(reverse('kardex:editar_stock_inicial', args=[movimiento.id]))

        self.assertEqual(response.status_code, 200)

    def test_pagina_productos_carga(self):
        user = User.objects.create_user(username='admin', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:productos_lista'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Productos')

    def test_crea_producto_desde_pagina_propia(self):
        user = User.objects.create_user(username='admin', password='admin12345')
        self.client.force_login(user)

        response = self.client.post(
            reverse('kardex:producto_crear'),
            data={
                'codigo_interno': 'PT-001',
                'nombre': 'Cafe oro',
                'categoria': 'Cafe',
                'tipo_producto': Producto.PRODUCTO_TERMINADO,
                'unidad_base': 'KG',
                'controla_stock': 'on',
                'afecta_kardex': 'on',
                'activo': 'on',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(Producto.objects.filter(codigo_interno='PT-001').exists())

    def test_edita_producto_desde_pagina_propia(self):
        user = User.objects.create_user(username='admin', password='admin12345')
        self.client.force_login(user)
        producto = Producto.objects.create(
            codigo_interno='PT-001',
            nombre='Cafe oro',
            tipo_producto=Producto.PRODUCTO_TERMINADO,
        )

        response = self.client.post(
            reverse('kardex:producto_editar', args=[producto.id]),
            data={
                'codigo_interno': 'PT-001',
                'nombre': 'Cafe oro exportable',
                'categoria': 'Cafe',
                'tipo_producto': Producto.PRODUCTO_TERMINADO,
                'unidad_base': 'KG',
                'controla_stock': 'on',
                'afecta_kardex': 'on',
                'activo': 'on',
            },
        )

        producto.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(producto.nombre, 'Cafe oro exportable')

    def test_pagina_documentos_carga_y_muestra_importados(self):
        importar_xml(FACTURA_XML, filename='factura.xml')
        user = User.objects.create_user(username='admin', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:documentos_lista'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '01-F001-123')
        self.assertContains(response, 'Proveedor SAC')

    def test_menu_documentos_apunta_a_pagina_propia(self):
        user = User.objects.create_user(username='admin', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:dashboard'))

        self.assertContains(response, reverse('kardex:documentos_lista'))

    def test_dashboard_muestra_analisis_compras_ventas_por_producto(self):
        producto, proveedor, cliente = self._crear_movimientos_dashboard()
        user = User.objects.create_user(username='dashboard', password='admin12345')
        self.client.force_login(user)

        response = self.client.get(reverse('kardex:dashboard'), {'producto': str(producto.id)})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Compras vs ventas')
        self.assertContains(response, 'Resumen por producto')
        self.assertContains(response, 'chartComprasVentasStock')
        self.assertContains(response, proveedor.razon_social)
        self.assertContains(response, cliente.razon_social)
        analisis = response.context['analisis']
        self.assertTrue(analisis.promedios_comparables)
        self.assertEqual(analisis.kpis['kg_comprados'], Decimal('20'))
        self.assertEqual(analisis.kpis['kg_vendidos'], Decimal('10'))
        self.assertEqual(analisis.kpis['stock_final'], Decimal('10'))
        self.assertEqual(analisis.kpis['precio_promedio_compra'], Decimal('10'))
        self.assertEqual(analisis.kpis['precio_promedio_venta'], Decimal('18'))
        self.assertEqual(analisis.kpis['margen_bruto_unitario'], Decimal('8'))
        self.assertEqual(analisis.kpis['margen_bruto_total'], Decimal('80'))
        self.assertIn('2026-06', analisis.periodos)

    def test_analisis_dashboard_no_mezcla_promedios_consolidados(self):
        self._crear_movimientos_dashboard()

        analisis = obtener_analisis_compras_ventas({})

        self.assertFalse(analisis.promedios_comparables)
        self.assertIsNone(analisis.kpis['precio_promedio_compra'])
        self.assertIsNone(analisis.kpis['precio_promedio_venta'])
        self.assertIsNone(analisis.kpis['margen_bruto_unitario'])
        self.assertEqual(len(analisis.resumen_productos), 1)
        self.assertEqual(analisis.resumen_productos[0]['margen_bruto_unitario'], Decimal('8'))

    def test_filtros_proveedor_y_cliente_aplican_a_compras_y_ventas(self):
        producto, proveedor, cliente = self._crear_movimientos_dashboard()

        analisis = obtener_analisis_compras_ventas(
            {
                'producto': str(producto.id),
                'proveedor': str(proveedor.id),
                'cliente': str(cliente.id),
            }
        )

        self.assertEqual(analisis.kpis['kg_comprados'], Decimal('20'))
        self.assertEqual(analisis.kpis['kg_vendidos'], Decimal('10'))
        self.assertEqual(analisis.kpis['margen_bruto_total'], Decimal('80'))

    def _crear_movimientos_dashboard(self):
        empresa = Entidad.objects.create(
            tipo_documento_identidad=Entidad.RUC,
            numero_documento='20999999999',
            razon_social='Empresa Principal SAC',
        )
        proveedor = Entidad.objects.create(
            tipo_documento_identidad=Entidad.RUC,
            numero_documento='20111111111',
            razon_social='Proveedor Dashboard SAC',
            es_proveedor=True,
        )
        cliente = Entidad.objects.create(
            tipo_documento_identidad=Entidad.RUC,
            numero_documento='20444444444',
            razon_social='Cliente Dashboard SAC',
            es_cliente=True,
        )
        producto = Producto.objects.create(
            codigo_interno='MP-DASH',
            nombre='Cafe dashboard',
            categoria='Cafe',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        compra = Documento.objects.create(
            tipo_documento=Documento.FACTURA,
            serie='F001',
            numero='900',
            fecha_emision=date(2026, 6, 1),
            entidad_emisor=proveedor,
            entidad_receptor=empresa,
            proveedor=proveedor,
            tipo_operacion=Documento.COMPRA,
            moneda='PEN',
            total=Decimal('236.00'),
            xml_hash='dashboard-compra',
            estado=Documento.CONFIRMADO,
        )
        DocumentoDetalle.objects.create(
            documento=compra,
            codigo_producto_xml='MP-DASH',
            descripcion_xml='Cafe dashboard',
            unidad_medida_xml='KG',
            cantidad=Decimal('20'),
            valor_unitario=Decimal('10'),
            precio_unitario=Decimal('11.80'),
            subtotal=Decimal('200'),
            igv=Decimal('36'),
            total=Decimal('236'),
            producto=producto,
            cantidad_base=Decimal('20'),
            afecta_kardex=True,
            estado_clasificacion=DocumentoDetalle.CLASIFICADO,
        )
        MovimientoKardex.objects.create(
            fecha=date(2026, 6, 1),
            producto=producto,
            documento_origen=compra,
            entidad=proveedor,
            tipo_movimiento=MovimientoKardex.ENTRADA,
            cantidad_entrada=Decimal('20'),
            costo_unitario_entrada=Decimal('10'),
            costo_total_entrada=Decimal('200'),
            stock_cantidad=Decimal('20'),
            stock_costo_unitario_promedio=Decimal('10'),
            stock_costo_total=Decimal('200'),
        )
        venta = Documento.objects.create(
            tipo_documento=Documento.FACTURA,
            serie='F002',
            numero='901',
            fecha_emision=date(2026, 6, 15),
            entidad_emisor=empresa,
            entidad_receptor=cliente,
            cliente=cliente,
            tipo_operacion=Documento.VENTA,
            moneda='PEN',
            total=Decimal('212.40'),
            xml_hash='dashboard-venta',
            estado=Documento.CONFIRMADO,
        )
        DocumentoDetalle.objects.create(
            documento=venta,
            codigo_producto_xml='MP-DASH',
            descripcion_xml='Cafe dashboard',
            unidad_medida_xml='KG',
            cantidad=Decimal('10'),
            valor_unitario=Decimal('18'),
            precio_unitario=Decimal('21.24'),
            subtotal=Decimal('180'),
            igv=Decimal('32.40'),
            total=Decimal('212.40'),
            producto=producto,
            cantidad_base=Decimal('10'),
            afecta_kardex=True,
            estado_clasificacion=DocumentoDetalle.CLASIFICADO,
        )
        MovimientoKardex.objects.create(
            fecha=date(2026, 6, 15),
            producto=producto,
            documento_origen=venta,
            entidad=cliente,
            tipo_movimiento=MovimientoKardex.SALIDA,
            cantidad_salida=Decimal('10'),
            costo_unitario_salida=Decimal('10'),
            costo_total_salida=Decimal('100'),
            stock_cantidad=Decimal('10'),
            stock_costo_unitario_promedio=Decimal('10'),
            stock_costo_total=Decimal('100'),
        )
        return producto, proveedor, cliente

    def test_formato_numero_usa_comas_y_dos_decimales(self):
        self.assertEqual(numero('1234567.891'), '1,234,567.89')
        self.assertEqual(numero(1000), '1,000.00')

    def test_formato_fecha_corta_usa_dia_mes_anio(self):
        self.assertEqual(fecha_corta(date(2026, 6, 25)), '25/06/2026')

    def test_formulario_stock_inicial_acepta_mas_de_dos_decimales(self):
        producto = Producto.objects.create(
            codigo_interno='MP-CAFE',
            nombre='Cafe pergamino',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        form = StockInicialForm(
            data={
                'fecha': '2026-06-25',
                'producto': producto.id,
                'cantidad': '1234.123456',
                'costo_unitario': '9.876543',
                'observacion': '',
            }
        )

        self.assertTrue(form.is_valid(), form.errors)

    def test_formularios_clasificacion_rechazan_factor_cero(self):
        producto = Producto.objects.create(
            codigo_interno='MP-FACTOR',
            nombre='Cafe factor',
            tipo_producto=Producto.MATERIA_PRIMA,
        )
        data = {
            'producto': producto.id,
            'factor_conversion': '0',
            'guardar_equivalencia': 'on',
        }

        detalle_form = ClasificarDetalleForm(data=data)
        bloque_form = ClasificarBloqueForm(data=data)

        self.assertFalse(detalle_form.is_valid())
        self.assertIn('factor_conversion', detalle_form.errors)
        self.assertFalse(bloque_form.is_valid())
        self.assertIn('factor_conversion', bloque_form.errors)
