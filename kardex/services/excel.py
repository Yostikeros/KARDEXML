import html
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse


EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
LEGACY_EXCEL_CONTENT_TYPE = "application/vnd.ms-excel"


def generar_excel_response(filename, sheet_title, headers, rows, metadata=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        return _generar_excel_html_response(filename.replace(".xlsx", ".xls"), headers, rows, metadata)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title[:31]

    current_row = 1
    if metadata:
        for label, value in metadata:
            worksheet.cell(row=current_row, column=1, value=label)
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            worksheet.cell(row=current_row, column=2, value=_excel_value(value))
            current_row += 1
        current_row += 1

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=current_row, column=column_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for row in rows:
        current_row += 1
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=current_row, column=column_index, value=_excel_value(value))
            if isinstance(value, (Decimal, int, float)):
                cell.number_format = "#,##0.00"
            elif isinstance(value, (date, datetime)):
                cell.number_format = "dd/mm/yyyy"

    _ajustar_columnas(worksheet, get_column_letter)

    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(buffer.getvalue(), content_type=EXCEL_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _excel_value(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    return value


def _ajustar_columnas(worksheet, get_column_letter):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 38)


def _generar_excel_html_response(filename, headers, rows, metadata=None):
    parts = [
        "<html><head><meta charset=\"utf-8\"></head><body><table border=\"1\">",
    ]
    if metadata:
        for label, value in metadata:
            parts.append(
                "<tr>"
                f"<td><strong>{html.escape(str(label))}</strong></td>"
                f"<td>{html.escape(str(_display_value(value)))}</td>"
                "</tr>"
            )
        parts.append(f"<tr><td colspan=\"{len(headers)}\"></td></tr>")

    parts.append("<tr>")
    for header in headers:
        parts.append(f"<th>{html.escape(str(header))}</th>")
    parts.append("</tr>")

    for row in rows:
        parts.append("<tr>")
        for value in row:
            parts.append(f"<td>{html.escape(str(_display_value(value)))}</td>")
        parts.append("</tr>")

    parts.append("</table></body></html>")
    response = HttpResponse("".join(parts), content_type=LEGACY_EXCEL_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _display_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, Decimal):
        return value
    return value
